import glob
import json
import os
import sys
from datetime import datetime
from zoneinfo import ZoneInfo

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(ROOT_DIR, ".env"))
LOCAL_TZ = ZoneInfo("Asia/Shanghai")


def truthy(value) -> bool:
    if value is True:
        return True
    if value is False or value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def safe_int(value, default=0) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return default


def load_manifest(results_dir: str) -> dict:
    path = os.path.join(results_dir, "manifest.json")
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as file:
            return json.load(file)
    except Exception:
        return {}


def target_date_text(manifest: dict) -> str:
    if isinstance(manifest, dict) and manifest.get("target_date"):
        return str(manifest["target_date"])[:10]
    return datetime.now(LOCAL_TZ).strftime("%Y-%m-%d")


def output_xlsx_path(results_dir: str, manifest: dict) -> str:
    configured = (os.getenv("OUTPUT_XLSX_PATH") or "").strip()
    if configured:
        return configured
    return os.path.join(results_dir, f"{target_date_text(manifest)}秒杀.xlsx")


def parse_account_line(line: str):
    if "----" in line:
        username, _ = line.split("----", 1)
    elif "," in line:
        username, _ = line.split(",", 1)
    else:
        return None
    username = username.strip()
    return username or None


def expected_accounts() -> tuple[dict[tuple[int, int], str], int]:
    lookup = {}
    for group_number in range(1, 2):
        raw = os.getenv(f"ACCOUNTS_BATCH{group_number}", "") or ""
        for account_index, line in enumerate(raw.splitlines(), start=1):
            username = parse_account_line(line.strip())
            if username:
                lookup[(group_number, account_index)] = username
    return lookup, len(lookup)


def find_json_files(results_dir: str) -> list[str]:
    paths = []
    for path in glob.glob(os.path.join(results_dir, "**", "*.json"), recursive=True):
        if os.path.basename(path).lower() == "manifest.json":
            continue
        if os.path.isfile(path):
            paths.append(path)
    return sorted(paths)


def seckill_record(row: dict) -> dict:
    activity = row.get("activity_records") if isinstance(row.get("activity_records"), dict) else {}
    records = activity.get("seckill")
    if isinstance(records, list) and records and isinstance(records[0], dict):
        return records[0]
    return {}


def cell_text(value, limit: int = 32000) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    else:
        text = str(value)
    text = " ".join(text.split())
    if len(text) <= limit:
        return text
    return text[:limit] + f"...(truncated,len={len(text)})"


def normalize_record(row: dict, payload: dict) -> dict:
    group_number = safe_int(row.get("group_number") or payload.get("group_number"), 0)
    account_index = safe_int(row.get("account_index"), 0)
    rec = seckill_record(row)
    diagnostics = row.get("runner_diagnostics") if isinstance(row.get("runner_diagnostics"), dict) else {}
    login_attempts = row.get("login_attempts") if isinstance(row.get("login_attempts"), list) else []
    cookie_count_raw = row.get("cookie_count") if row.get("cookie_count") is not None else rec.get("auth_cookie_count")
    success = truthy(row.get("sign_success")) or truthy(rec.get("success"))
    status = str(row.get("sign_status") or ("秒杀成功" if success else "秒杀失败")).strip()
    reason = str(row.get("detail_reason") or rec.get("last_response_message") or status).strip()
    return {
        "group_number": group_number,
        "account_index": account_index,
        "group_name": row.get("group_name") or payload.get("group_name") or f"seckill-batch{group_number}",
        "group_position": row.get("group_position") or f"{group_number}组账号{account_index}",
        "username": row.get("username") or row.get("masked_username") or f"账号{account_index}",
        "ga_job_started_at": diagnostics.get("ga_job_started_at") or "",
        "ga_run_seckill_started_at": diagnostics.get("ga_run_seckill_started_at") or "",
        "script_started_at": diagnostics.get("script_started_at") or "",
        "login_started_at": row.get("login_started_at") or (login_attempts[0].get("started_at") if login_attempts and isinstance(login_attempts[0], dict) else ""),
        "login_target": diagnostics.get("login_target") or "",
        "seckill_target": diagnostics.get("seckill_target") or "",
        "hard_stop_target": diagnostics.get("hard_stop_target") or "",
        "success": success,
        "skipped": truthy(row.get("seckill_skipped")),
        "status": status,
        "reason": reason,
        "password_error": truthy(row.get("password_error")),
        "risk_controlled": truthy(row.get("risk_controlled")),
        "sign_time": row.get("sign_time") or row.get("sign_completed_at") or "",
        "target_keyword": rec.get("target_keyword") or "",
        "title": rec.get("title") or "",
        "sku_code": rec.get("sku_code") or "",
        "goods_detail_access_id": rec.get("goods_detail_access_id") or "",
        "schedule_mode": rec.get("schedule_mode") or "",
        "calibration_source": rec.get("calibration_source") or "",
        "time_is_successes": rec.get("time_is_successes") or "",
        "attempts_sent": safe_int(rec.get("attempts_sent"), 0),
        "skipped_by_capacity": safe_int(rec.get("skipped_by_capacity"), 0),
        "median_rtt_ms": rec.get("median_rtt_ms"),
        "median_server_delta_ms": rec.get("median_server_delta_ms"),
        "success_sent_at": rec.get("success_sent_at") or "",
        "success_received_at": rec.get("success_received_at") or "",
        "token_extracted": truthy(row.get("token_extracted")),
        "secretkey_extracted": truthy(row.get("secretkey_extracted")),
        "m_site_token_bound": truthy(row.get("m_site_token_bound")) or truthy(rec.get("m_site_token_bound")),
        "cookie_count": safe_int(cookie_count_raw, 0),
        "cookie_attached_to_api": truthy(row.get("cookie_attached_to_api")) or truthy(rec.get("auth_cookie_attached")),
        "response_401_count": safe_int(rec.get("response_401_count"), 0),
        "response_401_ratio": rec.get("response_401_ratio") if rec.get("response_401_ratio") != "" else "",
        "auth_warning": rec.get("auth_warning") or "",
        "response_counts": rec.get("response_counts") if isinstance(rec.get("response_counts"), dict) else {},
        "first_response": rec.get("first_response") if isinstance(rec.get("first_response"), dict) else {},
        "first_401_response": rec.get("first_401_response") if isinstance(rec.get("first_401_response"), dict) else {},
        "first_non_401_response": rec.get("first_non_401_response") if isinstance(rec.get("first_non_401_response"), dict) else {},
        "success_response": rec.get("success_response") if isinstance(rec.get("success_response"), dict) else {},
    }


def load_results(results_dir: str) -> list[dict]:
    records = []
    for path in find_json_files(results_dir):
        try:
            with open(path, "r", encoding="utf-8") as file:
                payload = json.load(file)
        except Exception:
            continue
        rows = payload.get("results") if isinstance(payload, dict) else None
        if not isinstance(rows, list):
            continue
        for row in rows:
            if isinstance(row, dict):
                records.append(normalize_record(row, payload))
    return records


def missing_record(group_number: int, account_index: int, username: str) -> dict:
    return {
        "group_number": group_number,
        "account_index": account_index,
        "group_name": f"seckill-batch{group_number}",
        "group_position": f"{group_number}组账号{account_index}",
        "username": username,
        "ga_job_started_at": "",
        "ga_run_seckill_started_at": "",
        "script_started_at": "",
        "login_started_at": "",
        "login_target": "",
        "seckill_target": "",
        "hard_stop_target": "",
        "success": False,
        "skipped": False,
        "status": "未回传结果",
        "reason": "10:05 汇总时未下载到该账号结果",
        "password_error": False,
        "risk_controlled": False,
        "sign_time": "",
        "target_keyword": "",
        "title": "",
        "sku_code": "",
        "goods_detail_access_id": "",
        "schedule_mode": "",
        "calibration_source": "",
        "time_is_successes": "",
        "attempts_sent": 0,
        "skipped_by_capacity": 0,
        "median_rtt_ms": "",
        "median_server_delta_ms": "",
        "success_sent_at": "",
        "success_received_at": "",
        "token_extracted": False,
        "secretkey_extracted": False,
        "m_site_token_bound": False,
        "cookie_count": 0,
        "cookie_attached_to_api": False,
        "response_401_count": 0,
        "response_401_ratio": "",
        "auth_warning": "",
        "response_counts": {},
        "first_response": {},
        "first_401_response": {},
        "first_non_401_response": {},
        "success_response": {},
    }


def merge_expected(records: list[dict], lookup: dict[tuple[int, int], str]) -> list[dict]:
    by_key = {}
    extras = []
    for record in records:
        key = (safe_int(record.get("group_number"), 0), safe_int(record.get("account_index"), 0))
        if key[0] > 0 and key[1] > 0:
            by_key[key] = record
        else:
            extras.append(record)
    if not lookup:
        return list(by_key.values()) + extras
    merged = []
    for key in sorted(lookup):
        record = by_key.pop(key, None)
        if record:
            if not record.get("username"):
                record["username"] = lookup[key]
            merged.append(record)
        else:
            merged.append(missing_record(key[0], key[1], lookup[key]))
    merged.extend(by_key.values())
    merged.extend(extras)
    return merged


def sort_records(records: list[dict]) -> list[dict]:
    return sorted(
        records,
        key=lambda item: (
            2 if item.get("success") else (1 if item.get("skipped") else 0),
            safe_int(item.get("group_number"), 999),
            safe_int(item.get("account_index"), 999),
        ),
    )


def build_summary(records: list[dict], expected_total: int) -> dict:
    total = expected_total or len(records)
    success = sum(1 for item in records if item.get("success"))
    skipped = sum(1 for item in records if item.get("skipped"))
    failed = total - success - skipped
    by_goods = {}
    for item in records:
        if not item.get("success"):
            continue
        title = str(item.get("title") or "未知商品")
        by_goods[title] = by_goods.get(title, 0) + 1
    return {"total": total, "success": success, "skipped": skipped, "failed": failed, "by_goods": by_goods}


def build_message(records: list[dict], manifest: dict, expected_total: int) -> tuple[str, dict]:
    summary = build_summary(records, expected_total)
    lines = [
        f"{target_date_text(manifest)} 秒杀汇总",
        f"总账号: {summary['total']}",
        f"成功: {summary['success']}",
        f"配置关闭: {summary['skipped']}",
        f"失败/缺失: {summary['failed']}",
    ]
    if summary["by_goods"]:
        lines.append("成功商品:")
        for title, count in sorted(summary["by_goods"].items(), key=lambda item: (-item[1], item[0])):
            lines.append(f"- {title}: {count}")
    problem = [item for item in sort_records(records) if not item.get("success") and not item.get("skipped")]
    if problem:
        lines.append("异常账号:")
        for item in problem[:80]:
            lines.append(f"- {item['group_position']} {item['username']}: {item['reason']}")
        if len(problem) > 80:
            lines.append(f"...还有 {len(problem) - 80} 个异常账号，详见 XLSX")
    return "\n".join(lines), summary


def write_xlsx(path: str, records: list[dict]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "秒杀汇总"
    headers = [
        "序号",
        "组别",
        "账号",
        "状态",
        "原因",
        "目标关键词",
        "商品",
        "SKU",
        "商品ID",
        "调度模式",
        "发包数",
        "容量跳过数",
        "RTT中位数(ms)",
        "服务器时差中位数(ms)",
        "成功发送时间",
        "成功返回时间",
        "Runner到手时间",
        "秒杀命令开始",
        "脚本开始",
        "首次登录开始",
        "配置登录时间",
        "配置秒杀时间",
        "配置结束时间",
        "校准来源",
        "time.is成功次数",
        "token提取",
        "secretkey提取",
        "m站绑定",
        "Cookie数量",
        "Cookie已带入发包",
        "401次数",
        "401占比",
        "认证诊断",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    ok_font = Font(color="008000")
    skip_fill = PatternFill("solid", fgColor="FFF2CC")
    bad_fill = PatternFill("solid", fgColor="F8696B")
    bad_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for index, item in enumerate(sort_records(records), start=1):
        row = [
            index,
            item.get("group_position", ""),
            item.get("username", ""),
            "秒杀成功" if item.get("success") else item.get("status", "秒杀失败"),
            item.get("reason", ""),
            item.get("target_keyword", ""),
            item.get("title", ""),
            item.get("sku_code", ""),
            item.get("goods_detail_access_id", ""),
            item.get("schedule_mode", ""),
            item.get("attempts_sent", 0),
            item.get("skipped_by_capacity", 0),
            item.get("median_rtt_ms", ""),
            item.get("median_server_delta_ms", ""),
            item.get("success_sent_at", ""),
            item.get("success_received_at", ""),
            item.get("ga_job_started_at", ""),
            item.get("ga_run_seckill_started_at", ""),
            item.get("script_started_at", ""),
            item.get("login_started_at", ""),
            item.get("login_target", ""),
            item.get("seckill_target", ""),
            item.get("hard_stop_target", ""),
            item.get("calibration_source", ""),
            item.get("time_is_successes", ""),
            "是" if item.get("token_extracted") else "否",
            "是" if item.get("secretkey_extracted") else "否",
            "是" if item.get("m_site_token_bound") else "否",
            item.get("cookie_count", 0),
            "是" if item.get("cookie_attached_to_api") else "否",
            item.get("response_401_count", 0),
            item.get("response_401_ratio", ""),
            item.get("auth_warning", ""),
        ]
        sheet.append(row)
        row_index = sheet.max_row
        ratio_cell = sheet.cell(row_index, 32)
        if isinstance(item.get("response_401_ratio"), (int, float)):
            ratio_cell.number_format = "0.00%"
        for cell in sheet[row_index]:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        status_cell = sheet.cell(row_index, 4)
        if item.get("success"):
            status_cell.font = ok_font
        elif item.get("skipped"):
            status_cell.fill = skip_fill
        else:
            status_cell.fill = bad_fill
            status_cell.font = bad_font

    sheet.freeze_panes = "A2"
    widths = {
        "A": 8,
        "B": 16,
        "C": 24,
        "D": 14,
        "E": 42,
        "F": 16,
        "G": 30,
        "H": 18,
        "I": 34,
        "J": 14,
        "K": 12,
        "L": 12,
        "M": 16,
        "N": 22,
        "O": 28,
        "P": 28,
        "Q": 26,
        "R": 26,
        "S": 26,
        "T": 26,
        "U": 26,
        "V": 26,
        "W": 26,
        "X": 22,
        "Y": 16,
        "Z": 12,
        "AA": 16,
        "AB": 12,
        "AC": 12,
        "AD": 16,
        "AE": 12,
        "AF": 12,
        "AG": 42,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    stats_sheet = workbook.create_sheet("返回统计")
    stats_headers = [
        "序号",
        "组别",
        "账号",
        "商品",
        "发包数",
        "返回次数",
        "占比",
        "返回摘要",
        "首次返回",
        "首个401返回",
        "首个非401返回",
        "成功返回",
    ]
    stats_sheet.append(stats_headers)
    for cell in stats_sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    max_stats_rows = max(0, safe_int(os.getenv("XLSX_RESPONSE_STATS_MAX_ROWS"), 5000))
    stats_index = 0
    truncated = False
    for item in sort_records(records):
        counts = item.get("response_counts") if isinstance(item.get("response_counts"), dict) else {}
        attempts = max(0, safe_int(item.get("attempts_sent"), 0))
        if not counts:
            continue
        first_row_for_account = True
        for summary, count in sorted(counts.items(), key=lambda pair: (-safe_int(pair[1], 0), str(pair[0]))):
            if max_stats_rows and stats_index >= max_stats_rows:
                truncated = True
                break
            stats_index += 1
            response_count = safe_int(count, 0)
            ratio = response_count / attempts if attempts else ""
            stats_sheet.append(
                [
                    stats_index,
                    item.get("group_position", ""),
                    item.get("username", ""),
                    item.get("title", ""),
                    attempts,
                    response_count,
                    ratio,
                    cell_text(summary, 1200),
                    cell_text(item.get("first_response"), 1600) if first_row_for_account else "",
                    cell_text(item.get("first_401_response"), 1600) if first_row_for_account else "",
                    cell_text(item.get("first_non_401_response"), 1600) if first_row_for_account else "",
                    cell_text(item.get("success_response"), 1600) if first_row_for_account and item.get("success_response") else "",
                ]
            )
            first_row_for_account = False
            row_index = stats_sheet.max_row
            for cell in stats_sheet[row_index]:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(ratio, float):
                stats_sheet.cell(row_index, 7).number_format = "0.00%"
        if truncated:
            break
    if truncated:
        stats_sheet.append(
            [
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                f"返回统计超过 XLSX_RESPONSE_STATS_MAX_ROWS={max_stats_rows}，已截断；完整聚合仍保存在 result.json",
                "",
                "",
            ]
        )

    stats_sheet.freeze_panes = "A2"
    stats_widths = {
        "A": 8,
        "B": 16,
        "C": 24,
        "D": 30,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 90,
        "I": 70,
        "J": 70,
        "K": 70,
        "L": 70,
    }
    for column, width in stats_widths.items():
        stats_sheet.column_dimensions[column].width = width

    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    workbook.save(path)


def split_text(text: str, limit: int = 3900) -> list[str]:
    if len(text) <= limit:
        return [text]
    parts = []
    current = ""
    for line in text.splitlines(True):
        if len(current) + len(line) > limit and current:
            parts.append(current)
            current = ""
        current += line
    if current:
        parts.append(current)
    return parts


def send_telegram_message(text: str) -> bool:
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("[telegram] skipped text: token/chat_id empty")
        return False
    ok = True
    for part in split_text(text):
        try:
            response = requests.post(
                f"https://api.telegram.org/bot{token}/sendMessage",
                json={"chat_id": chat_id, "text": part},
                timeout=20,
            )
            if response.status_code != 200:
                print(f"[telegram] sendMessage failed: {response.status_code} {response.text[:300]}")
                ok = False
        except Exception as exc:
            print(f"[telegram] sendMessage exception: {type(exc).__name__}: {exc}")
            ok = False
    return ok


def send_telegram_document(path: str) -> bool:
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("[telegram] skipped xlsx: token/chat_id empty")
        return False
    if not os.path.exists(path):
        print(f"[telegram] xlsx not found: {path}")
        return False
    try:
        with open(path, "rb") as file:
            response = requests.post(
                f"https://api.telegram.org/bot{token}/sendDocument",
                data={"chat_id": chat_id},
                files={
                    "document": (
                        os.path.basename(path),
                        file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                timeout=60,
            )
        if response.status_code != 200:
            print(f"[telegram] sendDocument failed: {response.status_code} {response.text[:300]}")
            return False
        return True
    except Exception as exc:
        print(f"[telegram] sendDocument exception: {type(exc).__name__}: {exc}")
        return False


def main():
    results_dir = sys.argv[1] if len(sys.argv) > 1 else "results"
    manifest = load_manifest(results_dir)
    lookup, expected_total = expected_accounts()
    records = merge_expected(load_results(results_dir), lookup)
    message, summary = build_message(records, manifest, expected_total)
    xlsx_path = output_xlsx_path(results_dir, manifest)
    write_xlsx(xlsx_path, records)
    sent = False
    if (os.getenv("NOTIFY_CHANNELS") or "telegram").lower().find("telegram") >= 0:
        sent = send_telegram_message(message) or sent
        sent = send_telegram_document(xlsx_path) or sent
    print(message)
    print(
        f"[summary] total={summary['total']} success={summary['success']} "
        f"failed={summary['failed']} xlsx={xlsx_path} sent={'yes' if sent else 'no'}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
